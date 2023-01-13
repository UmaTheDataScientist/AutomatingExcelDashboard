#!/usr/bin/env python
# coding: utf-8

# In[1]:


#All Import Statements come here
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment  
from openpyxl.styles import PatternFill, Font


# In[2]:

def automate_excel_dashboard(file_name):
    #Insert your excel file name here
    #file_name = 'Bike_Sales_Playground.xlsx'


    # In[3]:


    #Reading the Excel file and the sheet name
    bike_df = pd.read_excel(file_name,sheet_name='bike_buyers')


    # In[4]:


    #Viewing the data
    bike_df.head()


    # # Creating a Working Sheet 

    # In[5]:


    #We don't want to mess with our raw data, thus, making a copy of it into a sheet called Working_Sheet.
    with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                   if_sheet_exists="replace" #Replacing the sheet if it already exists
                   ) as writer:  
    
        bike_df.to_excel(writer, sheet_name='Working_Sheet',index = False)#Setting index to False to avoid the unnecessary column Unnamed:0


    # In[6]:


    #Let's read the working sheet data into our dataframe
    bike_df = pd.read_excel(file_name,sheet_name='Working_Sheet')


    # In[7]:


    #Viewing the data
    bike_df.head()


    # # Data Cleaning

    # In[8]:


    #Dropping duplicates from the data
    bike_df.drop_duplicates(keep='first', inplace=True, ignore_index=False)


    # In[9]:


    #Replacing M to Married and S to Single in Marital Status column
    bike_df['Marital Status'] = bike_df['Marital Status'].replace('M','Married').replace('S','Single')


    # In[10]:


    #Replacing F to Female and M to Male in Gender column
    bike_df['Gender'] = bike_df['Gender'].replace('F','Female').replace('M','Male')


    # In[11]:


    #Viewing the changed column values
    bike_df.head()


    # In[12]:


    #Age is better in brackets
    #3 age brackets
    bike_df['Age brackets'] = bike_df['Age'].apply(lambda x: 'Less than 30' if x<=30 else('Greater than 55' if x>55 else '31 to 55'))


    # In[13]:


    bike_df['Age brackets'].unique()


    # In[14]:


    #Viewing the new column values
    bike_df.head()


    # In[15]:


    #Replacing Commute Distance value 10+ Miles to More than 10 Miles
    bike_df['Commute Distance'] = bike_df['Commute Distance'].replace('10+ Miles','More than 10 Miles')
    
    # In[16]:
    #Viewing the changed column values
    bike_df['Commute Distance'].unique()
    # In[17]:
    #Now that we have made all changes in the dataframe, let's load it into the excel file
    with pd.ExcelWriter(file_name,#Name of the Workbook
                    engine='openpyxl',#Name of the engine
                    mode='a',#Append mode
                   if_sheet_exists="replace" #Replacing the sheet if it already exists
                   ) as writer:  
    
        bike_df.to_excel(writer, sheet_name='Working_Sheet',index = False)#Setting index to False to avoid the unnecessary column Unnamed:0


    # # Exploratory Data Analysis

    # ### Pivot Table and Chart for Average Income Based on Gender and Purchase Data

    # In[18]:


    #Pivot table 1
    #Average Income per Gender based on Purchased Yes or No
    avg_gender_income_df = np.round(pd.pivot_table(bike_df,
                   values = 'Income',
                   index = ['Gender'],
                   columns = ['Purchased Bike'],
                   aggfunc = np.mean
                  ),2)
    
    
    # In[19]:
    
    
    avg_gender_income_df
    
    
    # In[20]:
    
    
    #Now that we have made all changes in the dataframe, let's load it into the excel file
    with pd.ExcelWriter(file_name,#Name of the Workbook
                        engine='openpyxl',#Name of the engine
                        mode='a',#Append mode
                       if_sheet_exists="replace" #Replacing the sheet if it already exists
                       ) as writer:  
        
        avg_gender_income_df.to_excel(writer, sheet_name='Average_Gender_Income')
    
    
    # In[21]:
    
    
    # loading workbook and selecting sheet
    wb = load_workbook(file_name)
    sheet = wb['Average_Gender_Income']
    
    
    # In[22]:
    
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Average Income by Gender and Purchase Data"
    chart1.y_axis.title = 'Gender'
    chart1.x_axis.title = 'Income'
    
    
    # In[23]:
    
    
    data1 = Reference(sheet, min_col=2, min_row=1, max_row=3, max_col=3)#Including Headers
    cats1 = Reference(sheet, min_col=1, min_row=2, max_row=3)#Not including headers
    chart1.add_data(data1, titles_from_data=True)
    chart1.dataLabels = DataLabelList() 
    chart1.dataLabels.showVal = True
    chart1.set_categories(cats1)
    chart1.shape = 4
    sheet.add_chart(chart1, "A10")
    wb.save(file_name)
    
    
    # ### Pivot Table and Chart for Customer Age Brackets with Purchase Data
    
    # In[24]:
    
    
    #Pivot table 2
    #Count of people purchasing bike per Age brackets
    count_age_brackets_purchase_df = pd.pivot_table(bike_df,
                   values = ['ID'],
                   index = 'Age brackets',
                   columns = 'Purchased Bike',
                   aggfunc = 'count'
                  )
    
    
    # In[25]:
    
    
    count_age_brackets_purchase_df
    
    
    # In[26]:
    
    
    count_age_brackets_purchase_df = count_age_brackets_purchase_df['ID']
    
    
    # In[27]:
    
    
    #Now that we have made all changes in the dataframe, let's load it into the excel file
    with pd.ExcelWriter(file_name,#Name of the Workbook
                        engine='openpyxl',#Name of the engine
                        mode='a',#Append mode
                       if_sheet_exists="replace" #Replacing the sheet if it already exists
                       ) as writer:  
        
        count_age_brackets_purchase_df.to_excel(writer, sheet_name='Count_Age_Brackets')
    
    
    # In[28]:
    
    
    # loading workbook and selecting sheet
    wb = load_workbook(file_name)
    sheet = wb['Count_Age_Brackets']
    
    
    # In[29]:
    
    
    chart2 = LineChart()
    chart2.style = 10
    chart2.title = "Count of Purchased with Age Brackets"
    chart2.y_axis.title = 'Count'
    chart2.x_axis.title = 'Age brackets'
    
    
    # In[30]:
    
    
    data2 = Reference(sheet, min_col=2, min_row=1, max_row=4, max_col=3)#Including Headers
    cats2 = Reference(sheet, min_col=1, min_row=2, max_row=4)#Not including headers
    chart2.add_data(data2, titles_from_data=True)
    chart2.dataLabels = DataLabelList() 
    chart2.dataLabels.showVal = True
    chart2.set_categories(cats2)
    chart2.shape = 4
    sheet.add_chart(chart2, "A10")
    wb.save(file_name)
    
    
    # ### Pivot Table and Chart for Commute Distance and Purchase Data
    
    # In[31]:
    
    
    #Pivot table 3
    #Count of people purchasing bike based on commute distance
    count_commute_distance_purchase_df = pd.pivot_table(bike_df,
                   values = ['ID'],
                   index = 'Commute Distance',
                   columns = 'Purchased Bike',
                   aggfunc = 'count'
                  )
    
    
    # In[32]:
    
    
    count_commute_distance_purchase_df
    
    
    # In[33]:
    
    
    count_commute_distance_purchase_df = count_commute_distance_purchase_df['ID']
    
    
    # In[34]:
    
    
    #Now that we have made all changes in the dataframe, let's load it into the excel file
    with pd.ExcelWriter(file_name,#Name of the Workbook
                        engine='openpyxl',#Name of the engine
                        mode='a',#Append mode
                       if_sheet_exists="replace" #Replacing the sheet if it already exists
                       ) as writer:  
        
        count_commute_distance_purchase_df.to_excel(writer, sheet_name='Count_Commute_Distance')
    
    
    # In[35]:
    
    
    # loading workbook and selecting sheet
    wb = load_workbook(file_name)
    sheet = wb['Count_Commute_Distance']
    
    
    # In[36]:
    
    
    chart3 = LineChart()
    chart3.style = 10
    chart3.title = "Count of Purchased with Commute Distance"
    chart3.y_axis.title = 'Count'
    chart3.x_axis.title = 'Commute Distance'
    
    
    # In[37]:
    
    
    data3 = Reference(sheet, min_col=2, min_row=1, max_row=6, max_col=3)#Including Headers
    cats3 = Reference(sheet, min_col=1, min_row=2, max_row=6)#Not including headers
    chart3.add_data(data3, titles_from_data=True)
    chart3.dataLabels = DataLabelList() 
    chart3.dataLabels.showVal = True
    chart3.set_categories(cats3)
    chart3.shape = 4
    sheet.add_chart(chart3, "A10")
    wb.save(file_name)
    
    
    # # Creating a Dashboard
    
    # In[38]:
    
    
    title_df = pd.DataFrame()
    
    
    # In[39]:
    #Now that we have made all changes in the dataframe, let's load it into the excel file
    with pd.ExcelWriter(file_name,#Name of the Workbook
                       engine='openpyxl',#Name of the engine
                       mode='a',#Append mode
                      if_sheet_exists="replace" #Replacing the sheet if it already exists
                      ) as writer:  
        
        title_df.to_excel(writer, sheet_name='Dashboard')
    
    # In[40]:
    # loading workbook and selecting sheet
    wb = load_workbook(file_name)
    sheet = wb['Dashboard']
    for x in range(1,22):
        sheet.merge_cells('A1:R4')
        
    cell = sheet.cell(row=1, column=1)  
    cell.value = 'Bike Sales Dashboard'  
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    cell.font  = Font(b=True, color="F8F8F8",size = 46)
    cell.fill = PatternFill("solid", fgColor="2591DB")
    
    
    # In[41]:
    
    
    #Adding all our pivot charts to the dashboard
    sheet.add_chart(chart1,'A5')
    sheet.add_chart(chart2,'J5')
    chart3.width = 31
    sheet.add_chart(chart3,'A20')
    wb.save(file_name)
    
